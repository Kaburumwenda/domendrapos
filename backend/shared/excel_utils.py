"""
Shared helpers for Excel import/export across apps.

Provides reusable openpyxl-based utilities for:
  - generating .xlsx workbooks from row data
  - parsing uploaded workbooks into normalized row dicts
  - generating styled templates with headers + example rows + instructions
"""
from io import BytesIO
from decimal import Decimal, InvalidOperation


# ─── Workbook builders ───────────────────────────────────────────────────────

def build_export_workbook(headers, rows):
    """Create a styled .xlsx workbook from a list of headers and row dicts.

    Args:
        headers: list of (label, field) tuples — label is the column header,
                 field is the key to look up in each row dict.
        rows: list of dicts containing the data.

    Returns: BytesIO positioned at 0 (ready for HttpResponse).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="FF1565C0", end_color="FF1565C0", fill_type="solid"
    )
    thin = Side(border_style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col_idx, (label, _field) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (_label, field) in enumerate(headers, 1):
            val = row_data.get(field)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if field in DEFAULT_NUMERIC_FIELDS:
                cell.alignment = Alignment(horizontal="right")

    # Auto-width
    for col_idx, (label, _f) in enumerate(headers, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(label)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                min_row=2, values_only=True):
            v = row[0]
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_template_workbook(headers, example_row, instructions, sheet_title="Data"):
    """Create a blank template .xlsx with headers, an example row, and an
    Instructions sheet.

    Args:
        headers: list of header labels (strings).
        example_row: list of values matching the headers (one example row).
        instructions: list of (text, is_heading) tuples.
        sheet_title: name for the main data sheet.

    Returns: BytesIO positioned at 0.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="FF1565C0", end_color="FF1565C0", fill_type="solid"
    )
    thin = Side(border_style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    ws.freeze_panes = "A2"

    # Example row
    for col_idx, val in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.border = border
        cell.font = Font(italic=True, color="FF757575")

    # Column auto-width
    for col_idx, label in enumerate(headers, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(len(label) + 4, 16)

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    for row_idx, (txt, bold) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=txt)
        if bold:
            cell.font = Font(bold=True, size=12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Workbook parser ─────────────────────────────────────────────────────────

def parse_workbook(wb, header_to_field, required_fields=None):
    """Parse an openpyxl workbook into a list of row dicts.

    Args:
        wb: openpyxl Workbook (already loaded).
        header_to_field: dict mapping normalized lowercased header → field name.
        required_fields: list of field names that must be present as columns.

    Returns: (rows, missing_required, skipped)
        - rows: list of dicts keyed by field name, each with '_row' for diagnostics.
        - missing_required: list of missing required field names.
        - skipped: count of empty rows skipped.
    """
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    if len(raw) < 2:
        return [], [], 0

    headers = [(str(h).strip().lower() if h is not None else "") for h in raw[0]]
    col_to_field = {}
    for col_idx, hdr in enumerate(headers):
        field = header_to_field.get(hdr)
        if field:
            col_to_field[col_idx] = field

    missing_required = []
    if required_fields:
        present = set(col_to_field.values())
        for f in required_fields:
            if f not in present:
                missing_required.append(f)

    rows = []
    skipped = 0
    for row_num, row in enumerate(raw[1:], 2):
        if row is None or all(v in (None, "", "None") for v in row):
            skipped += 1
            continue
        entry = {"_row": row_num}
        for col_idx, value in enumerate(row):
            field = col_to_field.get(col_idx)
            if not field or value is None:
                continue
            entry[field] = value
        rows.append(entry)
    return rows, missing_required, skipped


# ─── Normalizers ─────────────────────────────────────────────────────────────

def normalize_bools(cleaned, bool_fields):
    """Convert Yes/No/True/False/1/0 strings to booleans for the given fields.

    Removes the field from cleaned if the value can't be parsed.
    """
    for bool_field in bool_fields:
        v = cleaned.get(bool_field)
        if v is None:
            continue
        if isinstance(v, bool):
            continue
        s = str(v).strip().lower()
        if s in ("yes", "y", "1", "true"):
            cleaned[bool_field] = True
        elif s in ("no", "n", "0", "false"):
            cleaned[bool_field] = False
        else:
            cleaned.pop(bool_field, None)


def normalize_decimals(cleaned, dec_fields):
    """Convert string/int values to Decimal for the given fields.

    Converts bool to 0. Silently skips values that can't be parsed.
    """
    for dec_field in dec_fields:
        v = cleaned.get(dec_field)
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            cleaned[dec_field] = 0
            continue
        try:
            cleaned[dec_field] = str(Decimal(str(v)))
        except (InvalidOperation, ValueError):
            pass


def normalize_ints(cleaned, int_fields):
    """Convert string/float values to int for the given fields."""
    for int_field in int_fields:
        v = cleaned.get(int_field)
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            cleaned[int_field] = 0
            continue
        try:
            cleaned[int_field] = int(float(str(v)))
        except (ValueError, TypeError):
            pass


# ─── HTTP response helpers ───────────────────────────────────────────────────

def excel_response(buf, filename):
    """Build an HttpResponse for an xlsx file from a BytesIO buffer."""
    from django.http import HttpResponse
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# Fields that should be right-aligned in export
DEFAULT_NUMERIC_FIELDS = {
    "cost_price", "retail_price", "wholesale_price", "tax_rate",
    "weight", "quantity_on_hand", "reorder_level",
    "credit_limit", "current_credit_balance", "loyalty_points",
    "lead_time_days", "minimum_order_value", "rating",
    "unit_cost", "line_total", "subtotal", "discount_total",
    "tax_total", "shipping_cost", "grand_total",
    "quantity_ordered", "quantity_received",
}
