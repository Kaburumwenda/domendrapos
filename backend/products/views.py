from io import BytesIO
from decimal import Decimal, InvalidOperation
import random
import string

from django.http import HttpResponse
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend

from .models import Category, Product, ProductVariant, PriceList, ProductPriceOverride, Unit, Brand
from .filters import ProductFilter
from .serializers import (
    CategorySerializer,
    ProductSerializer,
    ProductVariantSerializer,
    PriceListSerializer,
    ProductPriceOverrideSerializer,
    UnitSerializer,
    BrandSerializer,
)


def generate_unique_sku(existing_skus=None, prefix="PRD"):
    """Generate a unique SKU like 'PRD-AB12CD'. The frontend ProductModal uses
    the same 'PRD-NNNNN' style; we add letters to reduce collision probability.
    `existing_skus` is an optional set/collection for in-batch uniqueness checks.
    """
    pool = string.ascii_uppercase + string.digits
    existing_skus = set(existing_skus or [])
    for _ in range(200):
        suffix = "".join(random.choice(pool) for _ in range(6))
        sku = f"{prefix}-{suffix}"
        if sku not in existing_skus and not Product.objects.filter(sku=sku).exists():
            return sku
    # Extremely unlikely fallback
    suffix = "".join(random.choice(pool) for _ in range(10))
    return f"{prefix}-{suffix}"


# ──────────────────────────────────────────────────────────────────────────────
# Excel column definitions (friendly header → serializer field)
# ──────────────────────────────────────────────────────────────────────────────
EXCEL_HEADERS = [
    "SKU",
    "Barcode",
    "Name",
    "Description",
    "Category",
    "Type",
    "Cost Price",
    "Retail Price",
    "Wholesale Price",
    "Tax Rate (%)",
    "Unit",
    "Items Per Unit",
    "Weight",
    "Brand",
    "Manufacturer",
    "Qty On Hand",
    "Reorder Level",
    "Expiry Date",
    "Is Active",
    "Is Sellable",
    "Is Purchasable",
    "Track Inventory",
]

# Map normalized header → serializer field
HEADER_TO_FIELD = {
    "sku": "sku",
    "barcode": "barcode",
    "name": "name",
    "description": "description",
    "category": "category",
    "type": "product_type",
    "cost price": "cost_price",
    "retail price": "retail_price",
    "wholesale price": "wholesale_price",
    "tax rate": "tax_rate",
    "tax rate (%)": "tax_rate",
    "unit": "unit",
    "items per unit": "items_per_unit",
    "items/unit": "items_per_unit",
    "expiry date": "expiry_date",
    "expiry": "expiry_date",
    "weight": "weight",
    "brand": "brand",
    "manufacturer": "manufacturer",
    "qty on hand": "quantity_on_hand",
    "quantity on hand": "quantity_on_hand",
    "reorder level": "reorder_level",
    "is active": "is_active",
    "is sellable": "is_sellable",
    "is purchasable": "is_purchasable",
    "track inventory": "track_inventory",
}

# Friendly names for export (ordered)
EXPORT_FIELDS = [
    ("SKU", "sku"),
    ("Barcode", "barcode"),
    ("Name", "name"),
    ("Description", "description"),
    ("Category", "category_name"),
    ("Type", "product_type"),
    ("Cost Price", "cost_price"),
    ("Retail Price", "retail_price"),
    ("Wholesale Price", "wholesale_price"),
    ("Tax Rate (%)", "tax_rate"),
    ("Unit", "unit"),
    ("Items Per Unit", "items_per_unit"),
    ("Weight", "weight"),
    ("Dimensions", "dimensions"),
    ("Brand", "brand"),
    ("Manufacturer", "manufacturer"),
    ("Qty On Hand", "quantity_on_hand"),
    ("Reorder Level", "reorder_level"),
    ("Expiry Date", "expiry_date"),
    ("Is Active", "is_active"),
    ("Is Sellable", "is_sellable"),
    ("Is Purchasable", "is_purchasable"),
    ("Track Inventory", "track_inventory"),
]


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["is_active", "parent"]
    search_fields = ["name", "description"]


class UnitViewSet(viewsets.ModelViewSet):
    queryset = Unit.objects.all()
    serializer_class = UnitSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["is_active"]
    search_fields = ["name", "abbreviation", "description"]
    ordering_fields = ["name", "created_at"]

    DEFAULT_UNITS = [
        ("Each", "ea", "Individual item"),
        ("Piece", "pc", "Single piece"),
        ("Kilogram", "kg", "Weight in kilograms"),
        ("Gram", "g", "Weight in grams"),
        ("Liter", "L", "Volume in liters"),
        ("Milliliter", "ml", "Volume in milliliters"),
        ("Meter", "m", "Length in meters"),
        ("Centimeter", "cm", "Length in centimeters"),
        ("Box", "box", "Box of items"),
        ("Pack", "pack", "Pack of items"),
        ("Dozen", "dz", "Dozen (12 items)"),
        ("Pair", "pair", "Pair of items"),
        ("Set", "set", "Set of items"),
        ("Roll", "roll", "Roll of material"),
        ("Bottle", "btl", "Bottle"),
        ("Carton", "ctn", "Carton of items"),
    ]

    @action(detail=False, methods=["post"], url_path="seed")
    def seed_units(self, request):
        """Seed default units of measure. Skips units that already exist."""
        created = []
        skipped = []
        for name, abbr, desc in self.DEFAULT_UNITS:
            obj, was_created = Unit.objects.get_or_create(
                name=name,
                defaults={"abbreviation": abbr, "description": desc, "is_active": True},
            )
            if was_created:
                created.append(name)
            else:
                skipped.append(name)
        return Response(
            {
                "created": created,
                "skipped": skipped,
                "total": Unit.objects.count(),
            },
            status=status.HTTP_200_OK,
        )


class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["is_active"]
    search_fields = ["name", "description"]
    ordering_fields = ["name", "created_at"]


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.select_related("category", "default_supplier").all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = ProductFilter
    search_fields = ["sku", "barcode", "name", "brand", "manufacturer"]
    ordering_fields = ["name", "retail_price", "created_at"]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    # ────────────────────────────────────────────────────────────────────────
    # Excel Export — GET /api/products/export-excel/
    # ────────────────────────────────────────────────────────────────────────
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """Stream all matching products (respecting query filters) as .xlsx."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        qs = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Items"

        # ── Header row ──
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="FF1565C0", end_color="FF1565C0", fill_type="solid"
        )
        thin = Side(border_style="thin", color="FFD0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, (label, _field) in enumerate(EXPORT_FIELDS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        ws.freeze_panes = "A2"

        # ── Data rows ──
        # Use serializer.to_representation for computed `quantity_on_hand` /
        # `reorder_level` and `category_name` fields
        serializer = self.get_serializer(qs, many=True)
        for row_idx, row_data in enumerate(serializer.data, 2):
            for col_idx, (_label, field) in enumerate(EXPORT_FIELDS, 1):
                val = row_data.get(field)
                if isinstance(val, bool):
                    val = "Yes" if val else "No"
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                if field in ("cost_price", "retail_price", "wholesale_price",
                             "tax_rate", "weight", "quantity_on_hand",
                             "reorder_level", "items_per_unit", "expiry_date"):
                    cell.alignment = Alignment(horizontal="right")

        # ── Auto-width columns ──
        for col_idx, (label, _f) in enumerate(EXPORT_FIELDS, 1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = len(label)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                    min_row=2, values_only=True):
                v = row[0]
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            # Cap widths to keep the file readable
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        # ── Build response ──
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"products_export_{__import__('datetime').datetime.now():%Y%m%d_%H%M%S}.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ────────────────────────────────────────────────────────────────────────
    # Excel Template — GET /api/products/import-excel-template/
    # ────────────────────────────────────────────────────────────────────────
    @action(detail=False, methods=["get"], url_path="import-excel-template")
    def import_excel_template(self, request):
        """Download a blank .xlsx template with headers + example row."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Items"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="FF1565C0", end_color="FF1565C0", fill_type="solid"
        )
        thin = Side(border_style="thin", color="FFD0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, label in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        ws.freeze_panes = "A2"

        # Example row
        example = [
            "SKU-001",
            "6001234567890",
            "Example Product",
            "Brief description here",
            "Electronics",
            "physical",
            "120.00",
            "250.00",
            "200.00",
            "16",
            "box",
            "24",
            "0.5",
            "Sample Brand",
            "Acme Corp",
            "50",
            "10",
            "2026-12-31",
            "Yes",
            "Yes",
            "Yes",
            "Yes",
        ]
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.border = border
            cell.font = Font(italic=True, color="FF757575")

        # Column auto-width
        for col_idx, label in enumerate(EXCEL_HEADERS, 1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = max(len(label) + 4, 16)

        # Notes sheet
        ws2 = wb.create_sheet("Instructions")
        notes = [
            ("BULK IMPORT — STOCK ITEMS", True),
            ("", False),
            ("Required fields", True),
            ("SKU — must be unique. If a product with this SKU exists, it will be updated.", False),
            ("Name — product name (cannot be blank).", False),
            ("", False),
            ("Optional fields", True),
            ("Barcode — GTIN/EAN. Defaults to blank.", False),
            ("Description — free-form text.", False),
            ("Category — match an existing Category name (case-insensitive). If blank, no category is assigned.", False),
            ("Type — one of: physical, service, digital, bundle. Defaults to physical.", False),
            ("Cost / Retail / Wholesale Price — numeric (e.g. 120.50). Default 0.", False),
            ("Tax Rate (%) — numeric. E.g. 16 = 16%. Stored as decimal.", False),
            ("Unit — unit abbreviation (e.g. ea, kg, L, pc). Defaults to 'ea'.", False),
            ("Items Per Unit — number of individual pieces per unit (e.g. 24 for a carton, 12 for a dozen). Defaults to 1.", False),
            ("Weight / Dimensions — decimal / text.", False),
            ("Brand / Manufacturer — free text.", False),
            ("Qty On Hand / Reorder Level — numeric. Sets the HQ-branch stock.", False),
            ("Expiry Date — optional date (YYYY-MM-DD) for perishable goods. Leave blank if not applicable.", False),
            ("Is Active / Is Sellable / Is Purchasable / Track Inventory — Yes / No.", False),
            ("", False),
            ("Behavior", True),
            ("Existing SKU → updates the row (upsert).", False),
            ("New SKU → creates the product.", False),
            ("Invalid rows are skipped with a row-level error in the response summary.", False),
        ]
        for row_idx, (txt, bold) in enumerate(notes, 1):
            cell = ws2.cell(row=row_idx, column=1, value=txt)
            if bold:
                cell.font = Font(bold=True, size=12)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="products_import_template.xlsx"'
        return resp

    # ────────────────────────────────────────────────────────────────
    # Shared: parse a workbook into a list of normalized row dicts.
    # ────────────────────────────────────────────────────────────────
    def _parse_workbook(self, wb):
        """Return (rows, missing_required, skipped).

        ``rows`` is a list of dicts (one per data row) keyed by serializer
        field, each carrying ``_row`` for diagnostics. Category stays as the
        original string value so callers can resolve it later. Empty SKUs
        are left blank — the caller decides whether to auto-generate.
        """
        ws = wb.active
        raw = list(ws.iter_rows(values_only=True))
        if len(raw) < 2:
            return [], [], 0

        headers = [(str(h).strip().lower() if h is not None else "") for h in raw[0]]
        col_to_field = {}
        for col_idx, hdr in enumerate(headers):
            field = HEADER_TO_FIELD.get(hdr)
            if field:
                col_to_field[col_idx] = field

        missing_required = []
        if "name" not in col_to_field.values():
            missing_required.append("Name")

        rows = []
        skipped = 0
        for row_num, row in enumerate(raw[1:], 2):
            if row is None or all(v in (None, "", "None") for v in row):
                skipped += 1
                continue
            entry = {"_row": row_num}
            for col_idx, value in enumerate(row):
                field = col_to_field.get(col_idx)
                if not field or value is None:
                    continue
                entry[field] = value
            rows.append(entry)
        return rows, missing_required, skipped

    # ────────────────────────────────────────────────────────────
    # Shared: build a lookup map for unit normalization.
    # Keys are lowercased name AND lowercased abbreviation → abbreviation.
    # E.g. {"each": "ea", "ea": "ea", "kilogram": "kg", "kg": "kg"}
    # ────────────────────────────────────────────────────────────
    def _build_unit_map(self):
        unit_map = {}
        for u in Unit.objects.all():
            abbr = (u.abbreviation or "").strip()
            if not abbr:
                continue
            unit_map[u.name.lower()] = abbr
            unit_map[abbr.lower()] = abbr
        return unit_map

    # ────────────────────────────────────────────────────────────
    # Shared: normalize a single entry dict (booleans, decimals, category).
    # Returns (cleaned_dict, error_msg_list).
    # ────────────────────────────────────────────────────────────
    def _normalize_entry(self, entry, cat_map, unit_map=None):
        cleaned = {}
        errors = []
        for k, val in entry.items():
            if k.startswith("_"):
                continue
            if k == "category":
                if val is None or str(val).strip() == "":
                    cleaned["category"] = None
                else:
                    key = str(val).strip().lower()
                    cat = cat_map.get(key)
                    if cat:
                        cleaned["category"] = cat.id
                    else:
                        if isinstance(val, (int, float)) and (
                            Category.objects.filter(id=int(val)).exists()
                        ):
                            cleaned["category"] = int(val)
                        else:
                            errors.append(f"Unknown category '{val}'")
                continue
            if k == "unit" and unit_map:
                if val is None or str(val).strip() == "":
                    cleaned["unit"] = "each"
                else:
                    key = str(val).strip().lower()
                    abbr = unit_map.get(key)
                    cleaned["unit"] = abbr if abbr else str(val).strip()
                continue
            cleaned[k] = val

        for bool_field in ("is_active", "is_sellable", "is_purchasable", "track_inventory"):
            v = cleaned.get(bool_field)
            if v is None:
                continue
            if isinstance(v, bool):
                continue
            s = str(v).strip().lower()
            if s in ("yes", "y", "1", "true"):
                cleaned[bool_field] = True
            elif s in ("no", "n", "0", "false"):
                cleaned[bool_field] = False
            else:
                cleaned.pop(bool_field, None)

        for dec_field in ("cost_price", "retail_price", "wholesale_price", "tax_rate",
                          "weight", "quantity_on_hand", "reorder_level"):
            v = cleaned.get(dec_field)
            if v is None or v == "":
                continue
            if isinstance(v, bool):
                cleaned[dec_field] = 0
                continue
            try:
                cleaned[dec_field] = str(Decimal(str(v)))
            except (InvalidOperation, ValueError):
                pass

        # Items per unit → integer
        ipu = cleaned.get("items_per_unit")
        if ipu is not None and ipu != "":
            try:
                cleaned["items_per_unit"] = int(float(ipu))
            except (ValueError, TypeError):
                pass

        # Expiry date → ISO date string (YYYY-MM-DD)
        ed = cleaned.get("expiry_date")
        if ed is not None and ed != "":
            if isinstance(ed, __import__('datetime').date):
                cleaned["expiry_date"] = ed.isoformat()
            elif isinstance(ed, __import__('datetime').datetime):
                cleaned["expiry_date"] = ed.date().isoformat()
            else:
                s = str(ed).strip()
                if s:
                    try:
                        __import__('datetime').date.fromisoformat(s)
                        cleaned["expiry_date"] = s
                    except ValueError:
                        errors.append(f"Invalid expiry date '{s}' (use YYYY-MM-DD)")
                        cleaned.pop("expiry_date", None)
        return cleaned, errors

    # ────────────────────────────────────────────────────────────
    # Excel Import (one-shot) — POST /api/products/import-excel/
    # ────────────────────────────────────────────────────────────
    @action(detail=False, methods=["post"], url_path="import-excel")
    def import_excel(self, request):
        """Bulk-import products from an .xlsx file (multipart 'file').
        Empty SKUs are auto-generated. Existing SKUs are updated.
        """
        from openpyxl import load_workbook

        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response(
                {"detail": "No file provided. Use multipart upload with key 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not uploaded.name.lower().endswith((".xlsx", ".xlsm")):
            return Response(
                {"detail": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(uploaded, read_only=True, data_only=True)
        except Exception as exc:
            return Response(
                {"detail": f"Could not read workbook: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows, missing_required, skipped = self._parse_workbook(wb)
        wb.close()
        if missing_required:
            return Response(
                {"detail": f"Missing required header column(s): {', '.join(missing_required)}. "
                 "Download the template first and keep its header row intact."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cat_map = {c.name.lower(): c for c in Category.objects.all()}
        unit_map = self._build_unit_map()
        existing_skus = set(Product.objects.values_list("sku", flat=True))

        created = 0
        updated = 0
        failed = 0
        errors = []
        for entry in rows:
            row_num = entry.pop("_row")
            cleaned, errs = self._normalize_entry(entry, cat_map, unit_map)
            if errs:
                failed += 1
                errors.append({"row": row_num, "sku": cleaned.get("sku", ""),
                                "detail": "; ".join(errs)})
                continue
            sku_val = cleaned.get("sku")
            if not sku_val or not str(sku_val).strip():
                sku_val = generate_unique_sku(existing_skus)
                cleaned["sku"] = sku_val
            existing_skus.add(sku_val)

            existing = Product.objects.filter(sku=sku_val).first()
            if existing:
                serializer = ProductSerializer(existing, data=cleaned, partial=True)
            else:
                serializer = ProductSerializer(data=cleaned)
            if serializer and serializer.is_valid():
                try:
                    serializer.save()
                    if existing:
                        updated += 1
                    else:
                        created += 1
                    continue
                except Exception as exc:
                    failed += 1
                    errors.append({"row": row_num, "sku": sku_val,
                                   "detail": f"Save error: {exc}"})
            else:
                failed += 1
                msg = "; ".join(f"{k}: {' / '.join(v)}" for k, v in serializer.errors.items() if v)
                errors.append({"row": row_num, "sku": sku_val,
                               "detail": msg or "Validation failed"})

        return Response({
            "created": created, "updated": updated, "skipped": skipped,
            "failed": failed, "total_processed": created + updated + failed,
            "errors": errors[:200], "errors_truncated": len(errors) > 200,
        }, status=status.HTTP_200_OK)

    # ────────────────────────────────────────────────────────────
    # Excel Parse / Preview — POST /api/products/parse-excel/
    # ────────────────────────────────────────────────────────────
    @action(detail=False, methods=["post"], url_path="parse-excel")
    def parse_excel(self, request):
        """Parse an uploaded .xlsx and return normalized rows for frontend
        preview/editing (does NOT persist). Empty SKUs are auto-filled so
        the user sees the final values before saving.
        """
        from openpyxl import load_workbook

        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded.name.lower().endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = load_workbook(uploaded, read_only=True, data_only=True)
        except Exception as exc:
            return Response({"detail": f"Could not read workbook: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        rows, missing_required, skipped = self._parse_workbook(wb)
        wb.close()
        if missing_required:
            return Response(
                {"detail": f"Missing required header column(s): {', '.join(missing_required)}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cat_map = {c.name.lower(): c for c in Category.objects.all()}
        categories = [{"id": c.id, "name": c.name} for c in Category.objects.all()]
        units = [{"id": u.id, "name": u.name, "abbreviation": u.abbreviation}
                 for u in Unit.objects.all().order_by("name")]
        unit_map = self._build_unit_map()
        existing_skus = set(Product.objects.values_list("sku", flat=True))

        out_rows = []
        errors = []
        for entry in rows:
            cat_name_raw = entry.get("category")
            cleaned, errs = self._normalize_entry(dict(entry), cat_map, unit_map)
            sku_val = cleaned.get("sku")
            if not sku_val or not str(sku_val).strip():
                sku_val = generate_unique_sku(existing_skus)
                cleaned["sku"] = sku_val
            existing_skus.add(sku_val)
            cleaned["_row"] = entry.get("_row")
            cleaned["_category_name"] = (str(cat_name_raw).strip()
                                        if cat_name_raw is not None and str(cat_name_raw).strip()
                                        else "")
            out_rows.append(cleaned)
            for e in errs:
                errors.append({"row": entry.get("_row"), "sku": cleaned.get("sku", ""), "detail": e})

        return Response({
            "rows": out_rows,
            "categories": categories,
            "product_types": [c[0] for c in Product.PRODUCT_TYPES],
            "units": units,
            "skipped": skipped,
            "errors": errors,
        }, status=status.HTTP_200_OK)

    # ────────────────────────────────────────────────────────────
    # Bulk Upsert (save preview) — POST /api/products/bulk-upsert/
    # ────────────────────────────────────────────────────────────
    @action(detail=False, methods=["post"], url_path="bulk-upsert")
    def bulk_upsert(self, request):
        """Save (possibly edited) rows from the preview. Empty SKUs are
        auto-generated. Existing SKUs are updated.
        """
        items = request.data.get("items") if isinstance(request.data, dict) else None
        if not items or not isinstance(items, list):
            return Response({"detail": "Body must be {\"items\": [...]}."}, status=status.HTTP_400_BAD_REQUEST)

        cat_map = {c.name.lower(): c for c in Category.objects.all()}
        unit_map = self._build_unit_map()
        existing_skus = set(Product.objects.values_list("sku", flat=True))

        created = 0
        updated = 0
        failed = 0
        errors = []
        for idx, item in enumerate(items, 1):
            cleaned, errs = self._normalize_entry(dict(item), cat_map, unit_map)
            if errs:
                failed += 1
                errors.append({"row": idx, "sku": cleaned.get("sku", ""),
                               "detail": "; ".join(errs)})
                continue
            sku_val = cleaned.get("sku")
            if not sku_val or not str(sku_val).strip():
                sku_val = generate_unique_sku(existing_skus)
                cleaned["sku"] = sku_val
            existing_skus.add(sku_val)

            existing = Product.objects.filter(sku=sku_val).first()
            if existing:
                serializer = ProductSerializer(existing, data=cleaned, partial=True)
            else:
                serializer = ProductSerializer(data=cleaned)
            if serializer and serializer.is_valid():
                try:
                    serializer.save()
                    if existing:
                        updated += 1
                    else:
                        created += 1
                    continue
                except Exception as exc:
                    failed += 1
                    errors.append({"row": idx, "sku": sku_val,
                                   "detail": f"Save error: {exc}"})
            else:
                failed += 1
                msg = "; ".join(f"{k}: {' / '.join(v)}" for k, v in serializer.errors.items() if v)
                errors.append({"row": idx, "sku": sku_val,
                               "detail": msg or "Validation failed"})

        return Response({
            "created": created, "updated": updated,
            "failed": failed, "total_processed": created + updated + failed,
            "errors": errors[:200], "errors_truncated": len(errors) > 200,
        }, status=status.HTTP_200_OK)


class ProductVariantViewSet(viewsets.ModelViewSet):
    queryset = ProductVariant.objects.all()
    serializer_class = ProductVariantSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["product"]


class PriceListViewSet(viewsets.ModelViewSet):
    queryset = PriceList.objects.all()
    serializer_class = PriceListSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["is_active"]


class ProductPriceOverrideViewSet(viewsets.ModelViewSet):
    queryset = ProductPriceOverride.objects.all()
    serializer_class = ProductPriceOverrideSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["product", "variant", "price_list", "branch"]

